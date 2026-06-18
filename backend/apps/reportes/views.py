import io
import csv
from django.views.generic import TemplateView
from django.db.models import Count, Avg, Max, Min, Q
from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.usuarios.decoradores import EvaluadorRequeridoMixin
from apps.usuarios.models import LogAccion
from apps.postulantes.models import Postulante, FichaSocioeconomica, SolicitudBeca
from apps.parametros.models import OpcionSocioeconomica


# ──────────────────────────────────────────────
# Dashboard estadístico
# ──────────────────────────────────────────────

class DashboardReportesView(EvaluadorRequeridoMixin, TemplateView):
    template_name = 'reportes/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Totales generales
        ctx['solicitudes_pendientes_count'] = SolicitudBeca.objects.filter(estado='Pendiente').count()
        ctx['solicitudes_hoy_count'] = Postulante.objects.filter(
            fecha_registro__date=timezone.now().date()
        ).count()
        
        total_p = Postulante.objects.count()
        completados_count = Postulante.objects.filter(ficha_completada=True).count()
        incompletos_count = total_p - completados_count
        
        ctx['total_postulantes'] = total_p
        ctx['completados_count'] = completados_count
        ctx['incompletos_count'] = incompletos_count
        
        if total_p > 0:
            ctx['porcentaje_completadas'] = round((completados_count / total_p) * 100, 1)
            ctx['porcentaje_incompletas'] = round((incompletos_count / total_p) * 100, 1)
        else:
            ctx['porcentaje_completadas'] = 0.0
            ctx['porcentaje_incompletas'] = 0.0

        ctx['completadas'] = SolicitudBeca.objects.filter(postulante__ficha_completada=True, rechazado=False).count()
        ctx['asignadas'] = SolicitudBeca.objects.filter(estado='Beca Asignada', rechazado=False).count()
        ctx['no_seleccionados'] = SolicitudBeca.objects.filter(estado='No seleccionado', rechazado=False).count()
        
        # Promedio de puntajes (Z) de todos los postulantes activos con ficha completa
        stats = SolicitudBeca.objects.filter(postulante__ficha_completada=True, rechazado=False).aggregate(
            avg=Avg('puntaje_total'),
            maximo=Max('puntaje_total'),
            minimo=Min('puntaje_total'),
        )
        ctx['promedio_z'] = round(stats['avg'] or 0, 2)
        ctx['puntaje_max'] = stats['maximo'] or 0
        ctx['puntaje_min'] = stats['minimo'] or 0

        # Para las tarjetas o compatibilidad
        ctx['total_evaluados'] = SolicitudBeca.objects.exclude(estado='Pendiente').filter(rechazado=False).count()
        ctx['total_aprobados'] = ctx['asignadas']
        ctx['total_rechazados'] = ctx['no_seleccionados']
        ctx['total_pendientes'] = ctx['completadas']

        # Distribución por rango de ingresos (estratos)
        opciones_ingresos = OpcionSocioeconomica.objects.filter(variable='rango_ingresos')
        estrato_data = []
        for opt in opciones_ingresos:
            rango_val = opt.opcion_texto
            total = Postulante.objects.filter(
                ficha_socioeconomica__rango_ingresos=rango_val
            ).count()
            aprobados = SolicitudBeca.objects.filter(
                postulante__ficha_socioeconomica__rango_ingresos=rango_val,
                estado='Beca Asignada',
                rechazado=False
            ).count()
            estrato_data.append({
                'estrato': opt.opcion_texto,
                'total': total,
                'aprobados': aprobados,
            })
        ctx['estrato_data'] = estrato_data

        # Distribución de estados (para gráfica de donut)
        ctx['estado_data'] = {
            'labels': ['Completada', 'Beca Asignada', 'No seleccionado'],
            'values': [
                ctx['completadas'],
                ctx['asignadas'],
                ctx['no_seleccionados'],
            ],
        }

        # Distribución por facultad
        facultad_query = SolicitudBeca.objects.filter(
            postulante__ficha_completada=True, 
            rechazado=False
        ).values('postulante__facultad').annotate(
            total=Count('id')
        ).order_by('-total')

        facultad_labels = []
        facultad_values = []
        for item in facultad_query:
            fac = item['postulante__facultad'] or 'Sin registrar'
            facultad_labels.append(fac)
            facultad_values.append(item['total'])
        
        ctx['facultad_data'] = {
            'labels': facultad_labels,
            'values': facultad_values,
        }

        return ctx


# ──────────────────────────────────────────────
# Exportar CSV
# ──────────────────────────────────────────────

class ExportarCSVView(EvaluadorRequeridoMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = (
            f'attachment; filename="postulantes_{timezone.now():%Y%m%d_%H%M}.csv"'
        )
        response.write('\ufeff')  # BOM para Excel en español

        writer = csv.writer(response)
        writer.writerow([
            'Cédula', 'Nombre Completo', 'Email', 'Carrera',
            'Celular', 'Dependencia', 'Rango Ingresos', 'Procedencia',
            'Puntaje Académico', 'Puntaje Socioeconómico', 'Puntaje Total',
            'Estado', 'Fecha Asignación',
        ])

        solicitudes = SolicitudBeca.objects.select_related(
            'postulante__user',
            'postulante__ficha_socioeconomica'
        ).order_by('-puntaje_total')

        for sol in solicitudes:
            p = sol.postulante
            user = p.user
            ficha = getattr(p, 'ficha_socioeconomica', None)
            
            carrera = p.carrera

            writer.writerow([
                p.cedula,
                user.get_full_name() or p.nombre_completo,
                user.email,
                carrera,
                p.telefono,
                ficha.dependencia if ficha else '',
                ficha.rango_ingresos if ficha else '',
                ficha.procedencia if ficha else '',
                sol.puntaje_academico or 0.0,
                sol.puntaje_socioeconomico or 0.0,
                sol.puntaje_total or 0.0,
                sol.estado,
                sol.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if sol.fecha_asignacion else '',
            ])

        LogAccion.objects.create(usuario=request.user, accion='exportar',
                                 detalles={'formato': 'CSV'})
        return response


# ──────────────────────────────────────────────
# Exportar Excel
# ──────────────────────────────────────────────

class ExportarExcelView(EvaluadorRequeridoMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Postulantes'

        # ── Estilos ────────────────────────────────
        azul = '1e3a5f'
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor=azul)
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            ('Cédula', 16), ('Nombre Completo', 30), ('Email', 28),
            ('Carrera', 25), ('Celular', 15),
            ('Dependencia', 20), ('Rango Ingresos', 25),
            ('Pts. Académico', 15), ('Pts. Socioecon.', 15), ('Pts. Total', 12),
            ('Estado', 18), ('Fecha Asignación', 20),
        ]

        # ── Cabecera ───────────────────────────────
        for col_idx, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 25

        # ── Datos ──────────────────────────────────
        solicitudes = SolicitudBeca.objects.select_related(
            'postulante__user',
            'postulante__ficha_socioeconomica'
        ).order_by('-puntaje_total')

        fill_aprobado = PatternFill('solid', fgColor='d5f5e3')
        fill_rechazado = PatternFill('solid', fgColor='fadbd8')
        fill_pendiente = PatternFill('solid', fgColor='fef9e7')

        for row_idx, sol in enumerate(solicitudes, 2):
            p = sol.postulante
            user = p.user
            ficha = getattr(p, 'ficha_socioeconomica', None)

            carrera = p.carrera
            estado = sol.estado
            
            fila_fill = (
                fill_aprobado if estado == 'Beca Asignada'
                else fill_rechazado if estado == 'No seleccionado'
                else fill_pendiente
            )

            row_data = [
                p.cedula, user.get_full_name() or p.nombre_completo, user.email,
                carrera, p.telefono,
                ficha.dependencia if ficha else '',
                ficha.rango_ingresos if ficha else '',
                float(sol.puntaje_academico) if sol.puntaje_academico is not None else 0.0,
                float(sol.puntaje_socioeconomico) if sol.puntaje_socioeconomico is not None else 0.0,
                float(sol.puntaje_total) if sol.puntaje_total is not None else 0.0,
                estado,
                sol.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if sol.fecha_asignacion else '',
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.fill = fila_fill
                if col_idx in (8, 9, 10):
                    cell.number_format = '#,##0.00'

        # ── Congelar cabecera ──────────────────────
        ws.freeze_panes = 'A2'

        # ── Hoja de estadísticas ───────────────────
        ws2 = wb.create_sheet('Estadísticas')
        ws2.append(['Métrica', 'Valor'])
        ws2.append(['Total postulantes', Postulante.objects.count()])
        ws2.append(['Beca Asignada', SolicitudBeca.objects.filter(estado='Beca Asignada').count()])
        ws2.append(['No seleccionado', SolicitudBeca.objects.filter(estado='No seleccionado').count()])
        ws2.append(['Postulación completada', SolicitudBeca.objects.filter(estado='Postulación completada').count()])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="becas_{timezone.now():%Y%m%d_%H%M}.xlsx"'
        )
        LogAccion.objects.create(usuario=request.user, accion='exportar',
                                 detalles={'formato': 'Excel'})
        return response
